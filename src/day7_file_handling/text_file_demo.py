# file=open("sample.txt","w")
# file.write("Hello, this is a file handling example")
# file.close()
# file=open("sample.txt","r")
# content=file.read()
# print(content)
# file.close()

# with open("sample.txt", "r") as file:
#     content=file.read()
#     print(content)


# try:
#     with open("missing.txt","r") as file:
#         print(file.read())
# except FileNotFoundError:
#     print("File not Found. Please check the filename.")


# import csv
# with open("D:\DS_AI_Internship\src\day7_file_handling\data1.csv","r") as file:
#     reader=csv.reader(file)
#     for row in reader:
#         print(row)


from openpyxl import load_workbook

file_path = r"D:\DS_AI_Internship\src\day7_file_handling\Book1.xlsx"

wb = load_workbook(file_path)
sheet = wb.active

for row in sheet.iter_rows(values_only=True):
    print(row)
